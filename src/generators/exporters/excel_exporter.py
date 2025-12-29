"""
Экспорт в Excel (.xlsx) через openpyxl.

Книга создаётся для ОДНОГО курса (экспорт вызывается в цикле по курсам в web-роуте).
Структура:
- Лист "Метаданные"
- Листы по типам вопросов: essay_gigachat, shortanswer, multichoice, matching, truefalse

Сильный фокус на читаемость:
- заголовки жирным + заливка
- перенос строк (wrap_text)
- границы
- адекватные ширины колонок
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


EXCEL_CELL_LIMIT = 32767


def _t(value: Any) -> str:
    return "" if value is None else str(value)


def _cell(value: Any) -> str:
    s = _t(value)
    if len(s) > EXCEL_CELL_LIMIT:
        return s[: EXCEL_CELL_LIMIT - 20] + " …(truncated)…"
    return s


def _header_text(metadata: Any, task_number: int) -> str:
    pk_prefix = _t(getattr(metadata, "pk_prefix", "ПК"))
    pk_id = _t(getattr(metadata, "pk_id", ""))
    ipk_prefix = _t(getattr(metadata, "ipk_prefix", "ИПК"))
    ipk_id = _t(getattr(metadata, "ipk_id", ""))
    description = _t(getattr(metadata, "description", ""))
    return f"- Задание {task_number} ({pk_prefix}-{pk_id} – {ipk_prefix}-{ipk_id} {description})"


class ExcelExporter:
    def export(self, questions, metadata, output_path: str, template_map: Dict[str, Any] | None = None) -> Dict[str, Any]:
        if not output_path or not str(output_path).strip():
            raise ValueError("output_path is empty")
        if questions is None:
            raise ValueError("questions is None")
        if not isinstance(questions, list):
            raise TypeError(f"questions must be a list, got {type(questions)!r}")
        if not questions:
            raise ValueError("Нет вопросов для экспорта.")

        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        # remove default sheet
        default_ws = wb.active
        wb.remove(default_ws)

        # Styles
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="E7E6E6")
        header_font = Font(bold=True)
        base_align = Alignment(vertical="top", wrap_text=True)

        # --- Metadata sheet ---
        meta_ws = wb.create_sheet("Метаданные")
        meta_rows = [
            ("Название документа", _t(getattr(metadata, "document_title", ""))),
            ("ПК", f"{_t(getattr(metadata,'pk_prefix','ПК'))}-{_t(getattr(metadata,'pk_id',''))}"),
            ("ИПК", f"{_t(getattr(metadata,'ipk_prefix','ИПК'))}-{_t(getattr(metadata,'ipk_id',''))}"),
            ("Описание", _t(getattr(metadata, "description", ""))),
        ]
        meta_ws.append(["Поле", "Значение"])
        for k, v in meta_rows:
            meta_ws.append([_cell(k), _cell(v)])

        self._style_table(meta_ws, header_row=1, border=border, header_fill=header_fill, header_font=header_font, align=base_align)
        self._set_widths(meta_ws, {1: 22, 2: 90})

        # --- Type sheets ---
        by_type: Dict[str, List[Any]] = {}
        for q in questions:
            by_type.setdefault(_t(getattr(q, "type", "")), []).append(q)

        rendered = 0
        skipped_types: Dict[str, int] = {}

        def _apply_cfg_widths(ws, q_type: str) -> None:
            cfg = (template_map or {}).get(q_type) if template_map else None
            if not isinstance(cfg, dict):
                return
            cols = ((cfg.get("layout") or {}).get(q_type, {}) or {}).get("table_cols_pct")
            if not (isinstance(cols, list) and all(isinstance(x, (int, float)) for x in cols)):
                return
            # Простейшая логика: берем первые N колонок листа и распределяем на 120 символов ширины.
            total = sum(cols) or 0
            if total <= 0:
                return
            max_col = min(len(cols), ws.max_column)
            for i in range(1, max_col + 1):
                pct = cols[i - 1] / total
                ws.column_dimensions[get_column_letter(i)].width = max(6, round(120 * pct, 1))

        for q_type in ["essay_gigachat", "shortanswer", "multichoice", "matching", "truefalse"]:
            items = by_type.get(q_type, [])
            ws = wb.create_sheet(q_type)

            if q_type == "essay_gigachat":
                ws.append(["№", "Заголовок", "Текст вопроса", "Эталонный ответ"])
                for idx, q in enumerate(items, start=1):
                    rendered += 1
                    ws.append(
                        [
                            idx,
                            _cell(_header_text(metadata, idx)),
                            _cell(getattr(q, "question_text", "")),
                            _cell(getattr(q, "reference_answer", "")),
                        ]
                    )
                self._set_widths(ws, {1: 5, 2: 45, 3: 80, 4: 80})

            elif q_type == "shortanswer":
                ws.append(["№", "Заголовок", "Текст вопроса", "Правильные ответы"])
                for idx, q in enumerate(items, start=1):
                    rendered += 1
                    answers = getattr(q, "correct_answers", None) or []
                    if not answers:
                        ref = _t(getattr(q, "reference_answer", ""))
                        answers = [ref] if ref else []
                    ws.append(
                        [
                            idx,
                            _cell(_header_text(metadata, idx)),
                            _cell(getattr(q, "question_text", "")),
                            _cell("\n".join(_t(a) for a in answers)),
                        ]
                    )
                self._set_widths(ws, {1: 5, 2: 45, 3: 80, 4: 60})

            elif q_type == "multichoice":
                ws.append(["№", "Заголовок", "Текст вопроса", "Варианты ответа", "Правильные ответы"])
                for idx, q in enumerate(items, start=1):
                    rendered += 1
                    answers = getattr(q, "answers", None) or []
                    correct = getattr(q, "correct_answers", None) or []
                    ws.append(
                        [
                            idx,
                            _cell(_header_text(metadata, idx)),
                            _cell(getattr(q, "question_text", "")),
                            _cell("\n".join(_t(a) for a in answers)),
                            _cell("\n".join(_t(a) for a in correct)),
                        ]
                    )
                self._set_widths(ws, {1: 5, 2: 45, 3: 70, 4: 60, 5: 50})

            elif q_type == "truefalse":
                ws.append(["№", "Заголовок", "Текст вопроса", "Варианты", "Правильный ответ"])
                for idx, q in enumerate(items, start=1):
                    rendered += 1
                    answers = getattr(q, "answers", None) or []
                    correct = getattr(q, "correct_answers", None) or []
                    ws.append(
                        [
                            idx,
                            _cell(_header_text(metadata, idx)),
                            _cell(getattr(q, "question_text", "")),
                            _cell("\n".join(_t(a) for a in answers)),
                            _cell("\n".join(_t(a) for a in correct)),
                        ]
                    )
                self._set_widths(ws, {1: 5, 2: 45, 3: 80, 4: 35, 5: 35})

            elif q_type == "matching":
                ws.append(["№", "Заголовок", "Текст вопроса", "Варианты ответа", "Пары (элемент → ответ)"])
                for idx, q in enumerate(items, start=1):
                    rendered += 1
                    answers = getattr(q, "matching_answers", None) or []
                    pairs = getattr(q, "matching_items", None) or []
                    pairs_text = "\n".join(
                        f"{_t(p.get('item',''))} → {_t(p.get('answer',''))}"
                        for p in pairs
                        if isinstance(p, dict)
                    )
                    ws.append(
                        [
                            idx,
                            _cell(_header_text(metadata, idx)),
                            _cell(getattr(q, "question_text", "")),
                            _cell("\n".join(_t(a) for a in answers)),
                            _cell(pairs_text),
                        ]
                    )
                self._set_widths(ws, {1: 5, 2: 45, 3: 60, 4: 45, 5: 70})

            self._style_table(ws, header_row=1, border=border, header_fill=header_fill, header_font=header_font, align=base_align)
            ws.freeze_panes = "A2"
            _apply_cfg_widths(ws, q_type)

        # Any other types -> sheet "other"
        other_types = [t for t in by_type.keys() if t not in {"essay_gigachat", "shortanswer", "multichoice", "matching", "truefalse"}]
        if other_types:
            ws = wb.create_sheet("other")
            ws.append(["Тип", "Название", "Текст вопроса"])
            for t in other_types:
                for q in by_type.get(t, []):
                    skipped_types[t] = skipped_types.get(t, 0) + 1
                    ws.append([_cell(t), _cell(getattr(q, "name", "")), _cell(getattr(q, "question_text", ""))])
            self._style_table(ws, header_row=1, border=border, header_fill=header_fill, header_font=header_font, align=base_align)
            self._set_widths(ws, {1: 18, 2: 30, 3: 100})
            ws.freeze_panes = "A2"

        wb.save(out)
        return {
            "output_path": str(out),
            "rendered_questions": rendered,
            "skipped_types": skipped_types,
        }

    def _style_table(
        self,
        ws,
        header_row: int,
        border: Border,
        header_fill: PatternFill,
        header_font: Font,
        align: Alignment,
    ) -> None:
        max_row = ws.max_row
        max_col = ws.max_column
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = align
                cell.border = border
                if r == header_row:
                    cell.fill = header_fill
                    cell.font = header_font

    def _set_widths(self, ws, widths: Dict[int, float]) -> None:
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width



